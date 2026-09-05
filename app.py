import re

import io
import html
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="SOLV — L2 TAT Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================================================
# STYLE
# ============================================================
st.markdown("""
<style>
    .main .block-container {max-width: 1450px; padding-top: 1.2rem;}
    .hero {
        background: linear-gradient(135deg,#173a5e 0%,#245f8f 55%,#2f75b5 100%);
        color:white; padding:28px 32px; border-radius:16px; margin-bottom:22px;
        box-shadow:0 8px 24px rgba(23,58,94,.15);
    }
    .hero h1 {margin:0 0 7px 0; font-size:32px; letter-spacing:.2px;}
    .hero p {margin:0; opacity:.9; font-size:15px;}
    .section {
        font-size:19px; font-weight:800; color:#173a5e;
        margin:28px 0 8px 0; padding-bottom:8px;
        border-bottom:2px solid #d9e2ec;
    }
    .smallnote {color:#64748b; font-size:13px; margin:5px 0 12px 0;}
    .kpi {
        color:white; border-radius:12px; padding:17px 18px;
        min-height:118px; box-shadow:0 4px 12px rgba(0,0,0,.08);
    }
    .kpi .label {font-size:12px; font-weight:800; letter-spacing:.2px;}
    .kpi .value {font-size:29px; font-weight:850; margin-top:10px;}
    .kpi .sub {font-size:11px; margin-top:6px; opacity:.9;}
    .blue {background:#2f75b5;} .green {background:#70ad47;}
    .orange {background:#ed7d31;} .red {background:#c00000;}
    .dark {background:#173a5e;}
</style>
""", unsafe_allow_html=True)

# ============================================================
# HELPERS
# ============================================================
def clean_text(s):
    return s.fillna("").astype(str).str.strip()

def parse_hms(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    s = str(value).strip()
    if not s or s.lower() in {"nan","nat","none","-"}:
        return np.nan
    try:
        if ":" in s:
            parts = s.split(":")
            if len(parts) == 3:
                h, m, sec = parts
                return float(h) + float(m)/60 + float(sec)/3600
            if len(parts) == 2:
                m, sec = parts
                return float(m)/60 + float(sec)/3600
        return float(s)
    except Exception:
        return np.nan

def format_hms(hours):
    if hours is None or pd.isna(hours):
        return "—"
    total = max(0, int(round(float(hours) * 3600)))
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    days = float(hours) / 24
    return f"{h}:{m:02d}:{s:02d} ({days:.1f} days)"

def format_days(hours):
    if hours is None or pd.isna(hours):
        return "—"
    return f"{float(hours)/24:.1f} days"

def unique_ticket_count(df):
    if df.empty:
        return 0
    return int(df["Ticket ID"].nunique())

def kpi(label, value, colour="blue", sub=""):
    st.markdown(
        f"""
        <div class="kpi {colour}">
          <div class="label">{html.escape(str(label))}</div>
          <div class="value">{html.escape(str(value))}</div>
          {f'<div class="sub">{html.escape(str(sub))}</div>' if sub else ''}
        </div>
        """,
        unsafe_allow_html=True
    )

def safe_pct(num, den):
    return (num / den * 100) if den else 0.0

# ============================================================
# DATA PREPARATION
# ============================================================
REQUIRED = ["Ticket ID", "Group", "Status", "Created time",
            "Closed time", "Resolution time (in hrs)"]


# ============================================================
# SOLV GROUP BUCKETS
# ============================================================
L2_GROUPS = {
    "ops - l2",
    "cd - seller support",
    "tech support",
    "logistics",
    "cd l2",
}

L1_GROUPS = {
    "cd l1 team",
    "no group",
}

# ============================================================
# LIVE TAT SOURCE — GOOGLE SHEET
# ============================================================
# The dashboard reads the current TAT sheet at runtime.
# Google Sheet from the user's screenshot:
# Spreadsheet ID: 1WONn8c8JQjmVYYYpt06jA-DUzH7YkroE4IxHoMO-TqM
# Sheet/tab: gid=0
#
# Expected columns:
#   Subcategory
#   SLA effective Jun'25
#
# If the sheet is private, add GOOGLE_SERVICE_ACCOUNT_JSON to
# Streamlit Secrets. If it is publicly readable, the CSV export
# works without credentials.

GOOGLE_TAT_SHEET_ID = "1WONn8c8JQjmVYYYpt06jA-DUzH7YkroE4IxHoMO-TqM"
GOOGLE_TAT_GID = "0"

def normalize_subcategory(value):
    if pd.isna(value):
        return ""
    s = str(value).strip().casefold()
    s = s.replace("&", " and ")
    s = re.sub(r"[-–—_/]+", " ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _read_tat_public_csv():
    import requests
    url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{GOOGLE_TAT_SHEET_ID}/export?format=csv&gid={GOOGLE_TAT_GID}"
    )
    r = requests.get(url, timeout=20)
    r.raise_for_status()
    if not r.text.strip():
        raise ValueError("The Google TAT sheet returned an empty response.")
    return pd.read_csv(io.StringIO(r.text))

def _read_tat_from_service_account():
    # Optional private-sheet support through Streamlit Secrets.
    import json
    import gspread
    from google.oauth2.service_account import Credentials

    secret = st.secrets.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not secret:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON not configured.")

    if isinstance(secret, str):
        info = json.loads(secret)
    else:
        info = dict(secret)

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(GOOGLE_TAT_SHEET_ID)
    ws = sh.get_worksheet_by_id(int(GOOGLE_TAT_GID))
    return pd.DataFrame(ws.get_all_records())

@st.cache_data(ttl=300, show_spinner=False)
def load_live_tat_sheet():
    errors = []

    # First try the public CSV export.
    try:
        raw_tat = _read_tat_public_csv()
    except Exception as e:
        errors.append(f"Public Google Sheet CSV export: {e}")
        raw_tat = None

    # If public export fails, try a service account configured in Secrets.
    if raw_tat is None:
        try:
            raw_tat = _read_tat_from_service_account()
        except Exception as e:
            errors.append(f"Service-account access: {e}")

    if raw_tat is None:
        raise RuntimeError(
            "Could not read the live TAT Google Sheet. "
            "Make the sheet accessible to the Streamlit app or configure "
            "GOOGLE_SERVICE_ACCOUNT_JSON in Streamlit Secrets. "
            + " | ".join(errors)
        )

    # Clean column names.
    raw_tat.columns = [str(c).strip() for c in raw_tat.columns]

    sub_col = next(
        (c for c in raw_tat.columns if normalize_subcategory(c) == "subcategory"),
        None
    )
    sla_col = next(
        (
            c for c in raw_tat.columns
            if normalize_subcategory(c) in {
                "sla effective jun25",
                "sla effective jun 25",
                "sla effective jun 2025",
            }
        ),
        None
    )

    if sub_col is None or sla_col is None:
        raise ValueError(
            "TAT sheet must contain 'Subcategory' and "
            "'SLA effective Jun'25' columns. "
            f"Found columns: {list(raw_tat.columns)}"
        )

    tat = raw_tat[[sub_col, sla_col]].copy()
    tat.columns = ["Subcategory", "TAT Raw"]

    tat["Subcategory"] = clean_text(tat["Subcategory"])
    tat["TAT Numeric"] = pd.to_numeric(
        tat["TAT Raw"].astype(str).str.strip().replace(
            {"NA": np.nan, "N/A": np.nan, "na": np.nan, "n/a": np.nan, "": np.nan}
        ),
        errors="coerce"
    )
    tat["_Key"] = tat["Subcategory"].apply(normalize_subcategory)

    # Ignore blank rows. If duplicate sub-categories exist, keep the last
    # populated numeric TAT and expose the duplicate count to the UI.
    tat = tat[tat["_Key"].ne("")].copy()

    duplicate_keys = (
        tat[tat["_Key"].duplicated(keep=False)]["_Key"]
        .value_counts()
    )
    duplicate_count = int((duplicate_keys > 1).sum())

    populated = tat[tat["TAT Numeric"].notna()].copy()
    tat_map = (
        populated.drop_duplicates("_Key", keep="last")
        .set_index("_Key")["TAT Numeric"]
        .to_dict()
    )

    # Keep the full mapping table exactly as represented in the source,
    # including NA TAT rows.
    display = tat[["Subcategory", "TAT Raw"]].copy()
    display["TAT (Days)"] = tat["TAT Numeric"].apply(
        lambda x: f"{int(x)} days" if pd.notna(x) else "NA"
    )

    return {
        "map": tat_map,
        "table": display,
        "rows": len(display),
        "mapped_rows": int(populated["_Key"].nunique()),
        "duplicate_subcategories": duplicate_count,
        "source_url": (
            f"https://docs.google.com/spreadsheets/d/"
            f"{GOOGLE_TAT_SHEET_ID}/edit?gid={GOOGLE_TAT_GID}"
        ),
    }

def get_tat_data():
    try:
        return load_live_tat_sheet(), None
    except Exception as e:
        return None, str(e)

def is_l2_group(series):
    return clean_text(series).str.casefold().isin(L2_GROUPS)

def is_l1_group(series):
    return clean_text(series).str.casefold().isin(L1_GROUPS)


def prepare_data(raw, tat_map):
    df = raw.copy()

    for c in ["Ticket ID","Group","Status","Created time","Closed time",
              "Resolution time (in hrs)","Reason for pendency",
              "Category","Sub-Category","Priority","Type",
              "Resolution status","First response time (in hrs)",
              "Agent","SO Helpline_L1","SO Helpline_L2"]:
        if c not in df.columns:
            df[c] = np.nan

    df["Ticket ID"] = df["Ticket ID"].astype(str).str.strip()
    df["_Group"] = clean_text(df["Group"]).replace("", "No Group")
    df["_Status"] = clean_text(df["Status"]).str.upper()

    df["_IsL2"] = df["_Group"].str.casefold().isin(L2_GROUPS)
    df["_IsL1"] = df["_Group"].str.casefold().isin(L1_GROUPS)

    # Treat Resolved as closed for resolution/SLA metrics because it has a
    # resolved state but may not have a Closed time.
    df["_ClosedLike"] = df["_Status"].isin(["CLOSED", "RESOLVED"])

    df["_Created"] = pd.to_datetime(df["Created time"], errors="coerce")
    df["_Closed"] = pd.to_datetime(df["Closed time"], errors="coerce")

    # Resolution time: prefer the raw export field.
    df["_ResolutionHours"] = df["Resolution time (in hrs)"].apply(parse_hms)

    # If a closed/resolved ticket has no usable raw resolution field,
    # derive elapsed Created -> Closed/Resolved where possible.
    missing_res = df["_ResolutionHours"].isna() & df["_ClosedLike"] & df["_Created"].notna() & df["_Closed"].notna()
    df.loc[missing_res, "_ResolutionHours"] = (
        (df.loc[missing_res, "_Closed"] - df.loc[missing_res, "_Created"])
        .dt.total_seconds() / 3600
    )

    df.loc[df["_ResolutionHours"] < 0, "_ResolutionHours"] = np.nan

    # Current age for tickets that are not closed/resolved.
    now = pd.Timestamp.now()
    open_like = ~df["_ClosedLike"]
    df["_CurrentAgeHours"] = np.nan
    age_mask = open_like & df["_Created"].notna()
    df.loc[age_mask, "_CurrentAgeHours"] = (
        (now - df.loc[age_mask, "_Created"]).dt.total_seconds() / 3600
    ).clip(lower=0)

    # Sub-category based TAT from the user's SLA sheet, not a fixed
    # 48-hour threshold.
    df["_SubCategoryKey"] = clean_text(df["Sub-Category"]).apply(normalize_subcategory)
    df["_TATDays"] = df["_SubCategoryKey"].map(tat_map)
    df["_TATHours"] = df["_TATDays"] * 24
    df["_TATStatus"] = "TAT Not Mapped"

    closed_valid_tat = (
        df["_ClosedLike"] &
        df["_ResolutionHours"].notna() &
        df["_TATHours"].notna()
    )
    df.loc[closed_valid_tat & (df["_ResolutionHours"] <= df["_TATHours"]), "_TATStatus"] = "Closed Within TAT"
    df.loc[closed_valid_tat & (df["_ResolutionHours"] > df["_TATHours"]), "_TATStatus"] = "Closed Beyond TAT"

    open_valid_tat = (
        open_like &
        df["_CurrentAgeHours"].notna() &
        df["_TATHours"].notna()
    )
    df.loc[open_valid_tat & (df["_CurrentAgeHours"] <= df["_TATHours"]), "_TATStatus"] = "Open Within TAT"
    df.loc[open_valid_tat & (df["_CurrentAgeHours"] > df["_TATHours"]), "_TATStatus"] = "Open Beyond TAT"

    # Backward-compatible label for downloads; dashboard logic now uses TAT.
    df["_SLA48"] = df["_TATStatus"]

    # Aging buckets for current open/pending workload.
    df["_AgingBucket"] = "Closed / Resolved"
    df.loc[open_like & (df["_CurrentAgeHours"] <= 24), "_AgingBucket"] = "0–24h"
    df.loc[open_like & (df["_CurrentAgeHours"] > 24) & (df["_CurrentAgeHours"] <= 48), "_AgingBucket"] = "24–48h"
    df.loc[open_like & (df["_CurrentAgeHours"] > 48) & (df["_CurrentAgeHours"] <= 72), "_AgingBucket"] = "48–72h"
    df.loc[open_like & (df["_CurrentAgeHours"] > 72) & (df["_CurrentAgeHours"] <= 168), "_AgingBucket"] = "3–7 days"
    df.loc[open_like & (df["_CurrentAgeHours"] > 168) & (df["_CurrentAgeHours"] <= 720), "_AgingBucket"] = "7–30 days"
    df.loc[open_like & (df["_CurrentAgeHours"] > 720) & (df["_CurrentAgeHours"] <= 2160), "_AgingBucket"] = "30–90 days"
    df.loc[open_like & (df["_CurrentAgeHours"] > 2160), "_AgingBucket"] = ">90 days"

    df["_CreatedDate"] = df["_Created"].dt.date
    df["_Month"] = df["_Created"].dt.strftime("%b %Y")
    df["_MonthSort"] = df["_Created"].dt.to_period("M").astype(str)
    df["_WeekStart"] = df["_Created"].dt.to_period("W-SUN").apply(lambda p: p.start_time if not pd.isna(p) else pd.NaT)
    df["_Week"] = np.where(
        df["_WeekStart"].notna(),
        df["_WeekStart"].dt.strftime("%d %b %Y"),
        ""
    )

    return df


def status_is_closed(df):
    return df["_ClosedLike"]


def cd_l1_analysis(data):
    """CD L1 intake and handoff analysis from the current raw Group snapshot."""
    l1 = data[
        data["_Group"].fillna("").astype(str).str.casefold().eq("cd l1 team")
    ].copy()

    total = unique_ticket_count(l1)

    closed = l1[l1["_ClosedLike"]]
    created = pd.to_datetime(l1["_Created"], errors="coerce")
    closed_time = pd.to_datetime(l1["_Closed"], errors="coerce")

    same_day = closed[
        created.loc[closed.index].dt.date.eq(
            closed_time.loc[closed.index].dt.date
        )
    ]

    # Current L2 is visible in the raw Group field, but the single
    # snapshot does not contain historical Group movement. Therefore this
    # metric is explicitly "currently in L2", not a proven movement count.
    l2_current = data[data["_IsL2"]]
    handoff_proxy = l1_l2_proxy_count(data)

    l1_open = l1[~l1["_ClosedLike"]]
    unassigned = clean_text(l1_open["Agent"]).str.casefold().isin(
        ["", "no agent", "not provided"]
    )
    fresh_unassigned = l1_open[
        unassigned &
        l1_open["_CurrentAgeHours"].notna() &
        (l1_open["_CurrentAgeHours"] < 48)
    ]

    phone_l1_closed = closed[
        clean_text(closed["Source"]).str.casefold().eq("phone") &
        created.loc[closed.index].dt.date.eq(
            closed_time.loc[closed.index].dt.date
        )
    ]

    return {
        "l1_raised": total,
        "l1_closed_same_day": unique_ticket_count(same_day),
        "l1_current_l2": unique_ticket_count(l2_current),
        "l1_l2_handoff_proxy": handoff_proxy,
        "l1_open": unique_ticket_count(l1_open),
        "l1_fresh_unassigned": unique_ticket_count(fresh_unassigned),
        "same_day_phone_proxy": unique_ticket_count(phone_l1_closed),
        "l1": l1,
        "fresh_unassigned": fresh_unassigned,
    }


def l1_agent_table(data):
    a = cd_l1_analysis(data)
    l1 = a["l1"]

    if l1.empty:
        return pd.DataFrame(columns=[
            "Agent","Tickets Raised","Closed / Resolved","Open / Pending",
            "Closed Same Day","Open <48h & Unassigned","Avg Resolution",
            "Max Resolution"
        ])

    rows = []
    for agent, g in l1.groupby(
        clean_text(l1["Agent"]).replace("", "No Agent"),
        sort=False
    ):
        r = g.loc[g["_ClosedLike"], "_ResolutionHours"].dropna()
        created = g["_Created"]
        closed = g["_Closed"]
        same_day_count = int(
            (
                g["_ClosedLike"] &
                created.notna() &
                closed.notna() &
                created.dt.date.eq(closed.dt.date)
            ).sum()
        )
        open_g = g[~g["_ClosedLike"]]
        age = open_g["_CurrentAgeHours"].dropna()
        fresh_unassigned = open_g[
            clean_text(open_g["Agent"]).str.casefold().isin(
                ["", "no agent", "not provided"]
            )
            & open_g["_CurrentAgeHours"].notna()
            & (open_g["_CurrentAgeHours"] < 48)
        ]

        rows.append({
            "Agent": agent,
            "Tickets Raised": unique_ticket_count(g),
            "Closed / Resolved": unique_ticket_count(g[g["_ClosedLike"]]),
            "Open / Pending": unique_ticket_count(open_g),
            "Closed Same Day": same_day_count,
            "Open <48h & Unassigned": unique_ticket_count(fresh_unassigned),
            "Avg Resolution": format_hms(r.mean()) if len(r) else "—",
            "Max Resolution": format_hms(r.max()) if len(r) else "—",
        })

    return pd.DataFrame(rows).sort_values(
        ["Open <48h & Unassigned", "Tickets Raised"],
        ascending=False
    ).reset_index(drop=True)


def fcr_l1_table(data):
    """
    FCR cannot be proven from this dump because there is no Call ID /
    contact-history field. Show a transparent same-day Phone proxy only.
    """
    a = cd_l1_analysis(data)
    l1 = a["l1"]

    if l1.empty:
        return pd.DataFrame(columns=[
            "Metric","Tickets","Definition"
        ])

    return pd.DataFrame([
        [
            "L1 tickets raised",
            a["l1_raised"],
            "Current Group = CD L1 Team / No Group"
        ],
        [
            "L1 closed same day",
            a["l1_closed_same_day"],
            "Created date = Closed date"
        ],
        [
            "Same-day Phone resolution proxy",
            a["same_day_phone_proxy"],
            "Phone source + closed on the same calendar day; not a proven FCR"
        ],
    ])


def l1_ticket_detail_table(data):
    a = cd_l1_analysis(data)
    x = a["fresh_unassigned"].copy()

    if x.empty:
        return pd.DataFrame(columns=[
            "Agent","Group","Ticket ID","Created Date",
            "Last Updated Date","Aging","Aging Days",
            "Category","Sub-Category","Status"
        ])

    out = pd.DataFrame()
    out["Agent"] = clean_text(x["Agent"]).replace("", "No Agent")
    out["Group"] = x["_Group"]
    out["Ticket ID"] = x["Ticket ID"]
    out["Created Date"] = x["_Created"].dt.strftime("%d %b %Y %H:%M:%S")
    out["Last Updated Date"] = pd.to_datetime(
        x["Last update time"], errors="coerce"
    ).dt.strftime("%d %b %Y %H:%M:%S")
    out["Aging"] = x["_CurrentAgeHours"].apply(format_hms)
    out["Aging Days"] = x["_CurrentAgeHours"].apply(
        lambda h: "—" if pd.isna(h) else f"{h/24:.1f} days"
    )
    out["Category"] = clean_text(x["Category"]).replace("", "(blank)")
    out["Sub-Category"] = clean_text(x["Sub-Category"]).replace("", "(blank)")
    out["Status"] = x["Status"]

    return out.sort_values(
        "Aging", ascending=False
    ).reset_index(drop=True)


def metrics(data):
    raised = unique_ticket_count(data)
    closed = unique_ticket_count(data[data["_ClosedLike"]])
    open_n = unique_ticket_count(data[~data["_ClosedLike"]])

    closed_valid = data[data["_ClosedLike"] & data["_ResolutionHours"].notna()]
    res = closed_valid["_ResolutionHours"]

    closed_within_tat = unique_ticket_count(
        data[data["_TATStatus"].eq("Closed Within TAT")]
    )
    closed_beyond_tat = unique_ticket_count(
        data[data["_TATStatus"].eq("Closed Beyond TAT")]
    )

    open_beyond_tat = unique_ticket_count(
        data[data["_TATStatus"].eq("Open Beyond TAT")]
    )
    tat_mapped = int(data["_TATHours"].notna().sum())
    tat_unmapped = int(data["_TATHours"].isna().sum())

    open_age = data.loc[~data["_ClosedLike"], "_CurrentAgeHours"].dropna()
    open_gt90 = int((open_age > 2160).sum())

    tat_closed_eligible = closed_within_tat + closed_beyond_tat

    return {
        "raised": raised,
        "closed": closed,
        "open": open_n,
        "closed_within_tat": closed_within_tat,
        "closed_beyond_tat": closed_beyond_tat,
        "tat_compliance_pct": safe_pct(closed_within_tat, tat_closed_eligible),
        "tat_closed_eligible": tat_closed_eligible,
        "open_beyond_tat": open_beyond_tat,
        "tat_mapped": tat_mapped,
        "tat_unmapped": tat_unmapped,
        "avg": res.mean() if len(res) else np.nan,
        "max_resolution": res.max() if len(res) else np.nan,
        "p90": res.quantile(0.90) if len(res) else np.nan,
        "p95": res.quantile(0.95) if len(res) else np.nan,
        "p99": res.quantile(0.99) if len(res) else np.nan,
        "open_gt90": open_gt90,
        "max_open_age": open_age.max() if len(open_age) else np.nan,
        "valid_resolution": len(res),
    }

def group_aging_table(data):
    rows = []
    for group, g in data.groupby("_Group", sort=False):
        age = g.loc[~g["_ClosedLike"], "_CurrentAgeHours"].dropna()
        rows.append({
            "Group": group,
            "Tickets": unique_ticket_count(g),
            "Open/Pending": unique_ticket_count(g[~g["_ClosedLike"]]),
            "Open Beyond TAT": unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Open >90 days": int((age > 2160).sum()),
            "Max Open Age": format_hms(age.max()) if len(age) else "—",
        })
    return pd.DataFrame(rows).sort_values("Tickets", ascending=False).reset_index(drop=True)

def subcategory_tat_performance_table(data):
    """Sub-category level TAT performance with the mapped TAT shown explicitly."""
    x = data.copy()
    x["_SubcatDisplay"] = clean_text(x["Sub-Category"]).replace("", "(blank)")
    rows=[]
    for sub,g in x.groupby("_SubcatDisplay", sort=False):
        raised = unique_ticket_count(g)
        closed = g[g["_ClosedLike"]]
        eligible_closed = closed[closed["_TATStatus"].isin(["Closed Within TAT","Closed Beyond TAT"])]
        within = unique_ticket_count(eligible_closed[eligible_closed["_TATStatus"].eq("Closed Within TAT")])
        beyond = unique_ticket_count(eligible_closed[eligible_closed["_TATStatus"].eq("Closed Beyond TAT")])
        open_beyond = unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")])
        tat_vals = g["_TATDays"].dropna().unique()
        tat_label = f"{int(tat_vals[0])} days" if len(tat_vals) else "Not mapped"
        rows.append({
            "Sub-Category": sub,
            "TAT": tat_label,
            "Tickets Raised": raised,
            "Closed / Resolved": unique_ticket_count(closed),
            "Closed Within TAT": within,
            "Closed Beyond TAT": beyond,
            "TAT Compliance %": round(safe_pct(within, within + beyond), 1) if (within + beyond) else np.nan,
            "Open / Pending": unique_ticket_count(g[~g["_ClosedLike"]]),
            "Open Beyond TAT": open_beyond,
        })
    return pd.DataFrame(rows).sort_values("Tickets Raised", ascending=False).reset_index(drop=True) if rows else pd.DataFrame(columns=[
        "Sub-Category","TAT","Tickets Raised","Closed / Resolved","Closed Within TAT","Closed Beyond TAT","TAT Compliance %","Open / Pending","Open Beyond TAT"
    ])


def open_reason_table(data):
    x = data[~data["_ClosedLike"]].copy()
    if x.empty:
        return pd.DataFrame(columns=["Open/Pending Reason","Tickets","% of Open/Pending"])

    x["_Reason"] = clean_text(x["Reason for pendency"])
    x.loc[x["_Reason"].eq(""), "_Reason"] = "Reason not provided"

    out = (
        x.groupby("_Reason")["Ticket ID"].nunique()
        .reset_index(name="Tickets")
        .rename(columns={"_Reason":"Open/Pending Reason"})
        .sort_values("Tickets", ascending=False)
        .reset_index(drop=True)
    )
    total = out["Tickets"].sum()
    out["% of Open/Pending"] = (out["Tickets"] / total * 100).round(1) if total else 0
    return out

def aging_table(data):
    order = ["0–24h","24–48h","48–72h","3–7 days","7–30 days","30–90 days",">90 days"]
    x = data[~data["_ClosedLike"]]
    out = x.groupby("_AgingBucket")["Ticket ID"].nunique().reindex(order, fill_value=0).reset_index()
    out.columns = ["Current Ticket Age","Tickets"]
    return out

def resolution_by_group(data):
    rows=[]
    for group,g in data.groupby("_Group",sort=False):
        r=g.loc[g["_ClosedLike"],"_ResolutionHours"].dropna()
        rows.append({
            "Group":group,
            "Closed/Resolved":unique_ticket_count(g[g["_ClosedLike"]]),
            "Closed Within TAT":unique_ticket_count(g[g["_TATStatus"].eq("Closed Within TAT")]),
            "Closed Beyond TAT":unique_ticket_count(g[g["_TATStatus"].eq("Closed Beyond TAT")]),
            "TAT Compliance %":round(
                safe_pct(
                    unique_ticket_count(g[g["_TATStatus"].eq("Closed Within TAT")]),
                    unique_ticket_count(g[g["_TATStatus"].isin(["Closed Within TAT","Closed Beyond TAT"])])
                ),1
            ),
            "Open Beyond TAT":unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Avg Resolution":format_hms(r.mean()) if len(r) else "—",
            "Max Resolution":format_hms(r.max()) if len(r) else "—",
        })
    return pd.DataFrame(rows).sort_values("Closed/Resolved",ascending=False).reset_index(drop=True)


def issue_breakup_table(data, column, denominator=None):
    x = data.copy()
    if column not in x.columns:
        return pd.DataFrame(columns=[column, "Tickets", "% of Tickets"])

    x["_Issue"] = clean_text(x[column])
    x.loc[x["_Issue"].eq(""), "_Issue"] = "(blank)"

    out = (
        x.groupby("_Issue")["Ticket ID"].nunique()
        .reset_index(name="Tickets")
        .rename(columns={"_Issue": column})
        .sort_values("Tickets", ascending=False)
        .reset_index(drop=True)
    )
    den = int(denominator if denominator is not None else unique_ticket_count(data))
    out["% of Tickets"] = (
        (out["Tickets"] / den * 100).round(1) if den else 0.0
    )
    return out


def category_subcategory_table(data):
    cols = ["Category", "Sub-Category"]
    x = data.copy()
    for c in cols:
        if c not in x.columns:
            x[c] = ""
        x[c] = clean_text(x[c]).replace("", "(blank)")

    out = (
        x.groupby(cols)["Ticket ID"]
        .nunique()
        .reset_index(name="Tickets")
        .sort_values("Tickets", ascending=False)
        .reset_index(drop=True)
    )
    total = unique_ticket_count(data)
    out["% of Tickets"] = (
        (out["Tickets"] / total * 100).round(1) if total else 0.0
    )
    return out


def other_group_pending_table(data):
    """Current tickets whose raw Group is not L2, with aging/SLA status, after the dashboard filters are applied."""
    x = data[~data["_IsL2"]].copy()
    if x.empty:
        return pd.DataFrame(columns=[
            "Current Group", "Tickets", "Open / Pending",
            "Open Beyond TAT", "Open >90 days", "Max Open Age"
        ])

    rows = []
    for group, g in x.groupby("_Group", sort=False):
        open_g = g[~g["_ClosedLike"]]
        ages = open_g["_CurrentAgeHours"].dropna()
        rows.append({
            "Current Group": group,
            "Tickets": unique_ticket_count(g),
            "Open / Pending": unique_ticket_count(open_g),
            "Open Beyond TAT": unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Open >90 days": int((ages > 2160).sum()),
            "Max Open Age": format_hms(ages.max()) if len(ages) else "—",
        })

    return (
        pd.DataFrame(rows)
        .sort_values(["Open / Pending", "Tickets"], ascending=False)
        .reset_index(drop=True)
    )



def other_group_open_breakup_table(data):
    """Open/pending tickets outside L2, broken down by current group and status."""
    x = data[~data["_IsL2"] & ~data["_ClosedLike"]].copy()
    cols = ["Current Group", "Open", "Pending", "Open / Pending", "Open Beyond TAT", "Agents"]
    if x.empty:
        return pd.DataFrame(columns=cols)

    x["_StatusLabel"] = x["_Status"].where(
        x["_Status"].isin(["OPEN", "PENDING"]), x["_Status"]
    )
    rows=[]
    for group,g in x.groupby("_Group", sort=False):
        agents = clean_text(g["Agent"]).replace("", "Not provided").unique().tolist()
        rows.append({
            "Current Group": group,
            "Open": unique_ticket_count(g[g["_Status"].eq("OPEN")]),
            "Pending": unique_ticket_count(g[g["_Status"].eq("PENDING")]),
            "Open / Pending": unique_ticket_count(g),
            "Open Beyond TAT": unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Agents": ", ".join(sorted(map(str, agents))),
        })
    return pd.DataFrame(rows, columns=cols).sort_values(
        ["Open / Pending", "Open Beyond TAT"], ascending=False
    ).reset_index(drop=True)


def other_group_open_agent_breakup_table(data):
    x = data[~data["_IsL2"] & ~data["_ClosedLike"]].copy()
    cols=["Current Group","Agent","Open","Pending","Open / Pending","Open Beyond TAT","Max Aging"]
    if x.empty:
        return pd.DataFrame(columns=cols)
    rows=[]
    for (group,agent),g in x.groupby(["_Group","Agent"], sort=False):
        agent_name = str(agent).strip() if str(agent).strip() else "Not provided"
        age=g["_CurrentAgeHours"].dropna()
        rows.append({
            "Current Group":group,
            "Agent":agent_name,
            "Open":unique_ticket_count(g[g["_Status"].eq("OPEN")]),
            "Pending":unique_ticket_count(g[g["_Status"].eq("PENDING")]),
            "Open / Pending":unique_ticket_count(g),
            "Open Beyond TAT":unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Max Aging":format_hms(age.max()) if len(age) else "—",
        })
    return pd.DataFrame(rows, columns=cols).sort_values(
        ["Open / Pending","Open Beyond TAT"], ascending=False
    ).reset_index(drop=True)


def other_group_open_category_breakup_table(data):
    x = data[~data["_IsL2"] & ~data["_ClosedLike"]].copy()
    cols=["Current Group","Category","Sub-Category","Open / Pending","Open Beyond TAT"]
    if x.empty:
        return pd.DataFrame(columns=cols)
    x["Category"] = clean_text(x["Category"]).replace("","(blank)")
    x["Sub-Category"] = clean_text(x["Sub-Category"]).replace("","(blank)")
    rows=[]
    for (group,cat,sub),g in x.groupby(["_Group","Category","Sub-Category"], sort=False):
        rows.append({
            "Current Group":group,
            "Category":cat,
            "Sub-Category":sub,
            "Open / Pending":unique_ticket_count(g),
            "Open Beyond TAT":unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
        })
    return pd.DataFrame(rows, columns=cols).sort_values(
        ["Open / Pending","Open Beyond TAT"], ascending=False
    ).reset_index(drop=True)


def other_group_agent_pending_table(data):
    """Agent-level view of currently open/pending tickets outside L2."""
    x = data[
        ~data["_IsL2"] &
        ~data["_ClosedLike"]
    ].copy()

    if x.empty:
        return pd.DataFrame(columns=[
            "Agent", "Group", "Tickets", "Open Beyond TAT",
            "Open >90 days", "Categories", "Sub-Categories",
            "Max Aging"
        ])

    rows = []
    for (agent, group), g in x.groupby(
        ["Agent", "_Group"], sort=False
    ):
        ages = g["_CurrentAgeHours"].dropna()
        cats = clean_text(g["Category"]).replace("", "(blank)").unique()
        subs = clean_text(g["Sub-Category"]).replace("", "(blank)").unique()

        rows.append({
            "Agent": agent if str(agent).strip() else "Not provided",
            "Group": group,
            "Tickets": unique_ticket_count(g),
            "Open Beyond TAT": unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Open >90 days": int((ages > 2160).sum()),
            "Categories": ", ".join(sorted(map(str, cats))),
            "Sub-Categories": ", ".join(sorted(map(str, subs))),
            "Max Aging": format_hms(ages.max()) if len(ages) else "—",
        })

    return pd.DataFrame(rows).sort_values(
        ["Open Beyond TAT", "Tickets"], ascending=False
    ).reset_index(drop=True)


def other_group_ticket_detail_table(data):
    """Ticket-level details for currently open/pending tickets outside L2."""
    x = data[
        ~data["_IsL2"] &
        ~data["_ClosedLike"]
    ].copy()

    if x.empty:
        return pd.DataFrame(columns=[
            "Agent", "Group", "Ticket ID", "Created Date",
            "Last Updated Date", "Aging", "Aging Days",
            "Category", "Sub-Category", "Open Reason"
        ])

    out = pd.DataFrame()
    out["Agent"] = clean_text(x["Agent"]).replace("", "Not provided")
    out["Group"] = x["_Group"]
    out["Ticket ID"] = x["Ticket ID"]
    out["Created Date"] = x["_Created"].dt.strftime("%d %b %Y %H:%M:%S")
    out["Last Updated Date"] = pd.to_datetime(
        x["Last update time"], errors="coerce"
    ).dt.strftime("%d %b %Y %H:%M:%S")
    out["Aging"] = x["_CurrentAgeHours"].apply(format_hms)
    out["Aging Days"] = x["_CurrentAgeHours"].apply(
        lambda h: "—" if pd.isna(h) else f"{h/24:.1f} days"
    )
    out["Category"] = clean_text(x["Category"]).replace("", "(blank)")
    out["Sub-Category"] = clean_text(x["Sub-Category"]).replace("", "(blank)")
    out["Open Reason"] = clean_text(x["Reason for pendency"]).replace(
        "", "Reason not provided"
    )

    out["_AgeSort"] = x["_CurrentAgeHours"].values
    return out.sort_values(
        ["Group", "Agent", "_AgeSort"], ascending=[True, True, False]
    ).drop(columns=["_AgeSort"]).reset_index(drop=True)


def week_breakup_table(data):
    rows = []
    for wk, g in data.groupby("_WeekStart", sort=True):
        mm = metrics(g)
        if pd.isna(wk):
            label = "Unknown"
        else:
            end = wk + pd.Timedelta(days=6)
            label = f"{wk.strftime('%d %b %Y')} - {end.strftime('%d %b %Y')}"
        rows.append({
            "Week": label,
            "Tickets Raised": mm["raised"],
            "Closed / Resolved": mm["closed"],
            "Open / Pending": mm["open"],
            "Closed Within TAT": mm["closed_within_tat"],
            "Closed Beyond TAT": mm["closed_beyond_tat"],
            "TAT Compliance %": round(mm["tat_compliance_pct"], 1),
            "Avg Resolution": format_hms(mm["avg"]),
            "Open Beyond TAT": mm["open_beyond_tat"],
            "Open >90 days": mm["open_gt90"],
        })
    return pd.DataFrame(rows)



def agent_ticket_detail_table(data):
    """Ticket-level ownership, aging and resolution detail for the selected population."""
    x = data.copy()

    def col(name):
        if name in x.columns:
            return clean_text(x[name]).replace("", "Not provided")
        return pd.Series(["Not provided"] * len(x), index=x.index)

    out = pd.DataFrame(index=x.index)
    out["Agent"] = col("Agent")
    out["Group"] = col("Group")
    out["Ticket ID"] = col("Ticket ID")

    created = pd.to_datetime(x["_Created"], errors="coerce")
    last_update = pd.to_datetime(x["Last update time"], errors="coerce")

    out["Created Date"] = created.dt.strftime("%d %b %Y %H:%M:%S")
    out["Last Updated Date"] = last_update.dt.strftime("%d %b %Y %H:%M:%S")

    # Current age for open/pending; resolution time for closed/resolved.
    out["Status"] = col("Status")
    out["Category"] = col("Category")
    out["Sub-Category"] = col("Sub-Category")
    out["Resolution Time"] = x["_ResolutionHours"].apply(format_hms)

    out["Current Aging"] = np.where(
        x["_ClosedLike"],
        x["_ResolutionHours"].apply(format_hms),
        x["_CurrentAgeHours"].apply(format_hms)
    )

    out["Aging Days"] = np.where(
        x["_ClosedLike"],
        x["_ResolutionHours"].apply(lambda h: "—" if pd.isna(h) else f"{h/24:.1f} days"),
        x["_CurrentAgeHours"].apply(lambda h: "—" if pd.isna(h) else f"{h/24:.1f} days")
    )

    out["TAT Days"] = x["_TATDays"].apply(
        lambda d: "—" if pd.isna(d) else f"{int(d)} days"
    )
    out["TAT Status"] = x["_TATStatus"]
    out["Open Reason"] = col("Reason for pendency")

    # Useful first-response metric when present in the dump.
    first_resp = x["First response time (in hrs)"].apply(parse_hms)
    out["First Response Time"] = first_resp.apply(format_hms)

    # Numeric helper for sorting/analysis.
    out["_SortAgeHours"] = np.where(
        x["_ClosedLike"],
        x["_ResolutionHours"],
        x["_CurrentAgeHours"]
    )

    return out.sort_values(
        ["Agent", "_SortAgeHours"],
        ascending=[True, False]
    ).drop(columns=["_SortAgeHours"]).reset_index(drop=True)


def agent_summary_table(data):
    detail = agent_ticket_detail_table(data)
    if detail.empty:
        return pd.DataFrame(columns=[
            "Agent", "Tickets", "Open / Pending", "Closed / Resolved",
            "Open Beyond TAT", "Avg Resolution", "Max Resolution"
        ])

    rows = []
    for agent, g in data.groupby(
        clean_text(data["Agent"]).replace("", "Not provided"),
        sort=False
    ):
        closed = g[g["_ClosedLike"]]
        open_g = g[~g["_ClosedLike"]]
        r = closed["_ResolutionHours"].dropna()
        age = open_g["_CurrentAgeHours"].dropna()
        rows.append({
            "Agent": agent,
            "Tickets": unique_ticket_count(g),
            "Open / Pending": unique_ticket_count(open_g),
            "Closed / Resolved": unique_ticket_count(closed),
            "Open Beyond TAT": unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Open >90 days": int((age > 2160).sum()),
            "Avg Resolution": format_hms(r.mean()) if len(r) else "—",
            "Max Resolution": format_hms(r.max()) if len(r) else "—",
            "Max Current Age": format_hms(age.max()) if len(age) else "—",
        })
    return pd.DataFrame(rows).sort_values(
        ["Open Beyond TAT", "Tickets"], ascending=False
    ).reset_index(drop=True)



def group_set_for_agent(data, agent):
    """Return all current groups handled by the selected agent."""
    if data is None or data.empty or "Agent" not in data.columns:
        return pd.DataFrame()

    x = data[
        clean_text(data["Agent"]).eq(agent)
    ].copy()

    return x


def agent_cross_group_summary(data, agent):
    """Current-group workload for a selected agent across the full dump."""
    x = group_set_for_agent(data, agent)

    if x.empty:
        return pd.DataFrame(columns=[
            "Group", "Tickets", "Open / Pending", "Closed / Resolved",
            "Open Beyond TAT", "Open >90 days", "Avg Resolution",
            "Max Open Aging"
        ])

    rows = []
    for group, g in x.groupby("_Group", sort=False):
        open_g = g[~g["_ClosedLike"]]
        closed_g = g[g["_ClosedLike"]]
        res = closed_g["_ResolutionHours"].dropna()
        age = open_g["_CurrentAgeHours"].dropna()

        rows.append({
            "Group": group,
            "Tickets": unique_ticket_count(g),
            "Open / Pending": unique_ticket_count(open_g),
            "Closed / Resolved": unique_ticket_count(closed_g),
            "Open Beyond TAT": unique_ticket_count(g[g["_TATStatus"].eq("Open Beyond TAT")]),
            "Open >90 days": int((age > 2160).sum()),
            "Avg Resolution": format_hms(res.mean()) if len(res) else "—",
            "Max Open Aging": format_hms(age.max()) if len(age) else "—",
        })

    return pd.DataFrame(rows).sort_values(
        ["Open Beyond TAT", "Tickets"],
        ascending=False
    ).reset_index(drop=True)


def agent_cross_group_tickets(data, agent, selected_group=None):
    """Ticket-level cross-group detail for the selected agent."""
    x = group_set_for_agent(data, agent)

    if selected_group and selected_group != "All":
        x = x[x["_Group"].eq(selected_group)]

    if x.empty:
        return pd.DataFrame(columns=[
            "Group", "Ticket ID", "Status", "Created Date",
            "Last Updated Date", "Category", "Sub-Category",
            "Resolution Time", "Current Aging", "Aging Days",
            "TAT Status", "Open Reason"
        ])

    out = pd.DataFrame()
    out["Group"] = x["_Group"]
    out["Ticket ID"] = x["Ticket ID"]
    out["Status"] = x["Status"]
    out["Created Date"] = x["_Created"].dt.strftime("%d %b %Y %H:%M:%S")
    out["Last Updated Date"] = pd.to_datetime(
        x["Last update time"], errors="coerce"
    ).dt.strftime("%d %b %Y %H:%M:%S")
    out["Category"] = clean_text(x["Category"]).replace("", "(blank)")
    out["Sub-Category"] = clean_text(x["Sub-Category"]).replace("", "(blank)")
    out["Resolution Time"] = x["_ResolutionHours"].apply(format_hms)

    out["Current Aging"] = np.where(
        x["_ClosedLike"],
        x["_ResolutionHours"].apply(format_hms),
        x["_CurrentAgeHours"].apply(format_hms)
    )

    out["Aging Days"] = np.where(
        x["_ClosedLike"],
        x["_ResolutionHours"].apply(
            lambda h: "—" if pd.isna(h) else f"{h/24:.1f} days"
        ),
        x["_CurrentAgeHours"].apply(
            lambda h: "—" if pd.isna(h) else f"{h/24:.1f} days"
        )
    )

    out["TAT Days"] = x["_TATDays"].apply(
        lambda d: "—" if pd.isna(d) else f"{int(d)} days"
    )
    out["TAT Status"] = x["_TATStatus"]
    out["Open Reason"] = clean_text(
        x["Reason for pendency"]
    ).replace("", "Reason not provided")

    out["_sort_age"] = np.where(
        x["_ClosedLike"],
        x["_ResolutionHours"],
        x["_CurrentAgeHours"]
    )

    return out.sort_values(
        "_sort_age",
        ascending=False
    ).drop(columns="_sort_age").reset_index(drop=True)


def l1_l2_proxy_count(data):
    """Unique current L2 tickets whose CreatedBy also appears in current L1.

    This is a proxy only because the raw dump has current Group, not historical
    Group movement/transfer history.
    """
    l1 = data[data["_IsL1"]].copy()
    l2 = data[data["_IsL2"]].copy()
    if l1.empty or l2.empty:
        return 0

    l1_keys = clean_text(l1["CreatedBy"]).str.casefold()
    valid = set(l1_keys[l1_keys.ne("")].unique())
    l2_keys = clean_text(l2["CreatedBy"]).str.casefold()
    return unique_ticket_count(l2[l2_keys.isin(valid)])


def l1_created_to_l2_table(data):
    """Current-state L1 -> L2 handoff proxy, broken down by CreatedBy."""
    l1 = data[data["_IsL1"]].copy()
    l2 = data[data["_IsL2"]].copy()
    cols = [
        "L1 Champ / Created By", "L1 Tickets Raised",
        "Current L2 Tickets Created By Champ", "L2 Groups",
        "L2 Open / Pending", "L2 Closed / Resolved"
    ]
    if l1.empty or l2.empty:
        return pd.DataFrame(columns=cols)

    l1["_CreatorKey"] = clean_text(l1["CreatedBy"]).str.casefold()
    l2["_CreatorKey"] = clean_text(l2["CreatedBy"]).str.casefold()
    valid = set(l1.loc[l1["_CreatorKey"].ne(""), "_CreatorKey"].unique())
    l2 = l2[l2["_CreatorKey"].isin(valid)].copy()
    if l2.empty:
        return pd.DataFrame(columns=cols)

    counts = (
        l1[l1["_CreatorKey"].isin(valid)]
        .groupby("_CreatorKey")["Ticket ID"].nunique()
        .reset_index(name="L1 Tickets Raised")
    )
    display_map = {}
    for _, r in l1.iterrows():
        k = r["_CreatorKey"]
        v = str(r["CreatedBy"]).strip()
        if k and k not in display_map and v:
            display_map[k] = v

    rows=[]
    for _, r in counts.iterrows():
        key = r["_CreatorKey"]
        g = l2[l2["_CreatorKey"].eq(key)]
        if g.empty:
            continue
        op = g[~g["_ClosedLike"]]
        cl = g[g["_ClosedLike"]]
        rows.append({
            "L1 Champ / Created By": display_map.get(key, key),
            "L1 Tickets Raised": int(r["L1 Tickets Raised"]),
            "Current L2 Tickets Created By Champ": unique_ticket_count(g),
            "L2 Groups": ", ".join(sorted(clean_text(g["_Group"]).unique().tolist())),
            "L2 Open / Pending": unique_ticket_count(op),
            "L2 Closed / Resolved": unique_ticket_count(cl),
        })
    return pd.DataFrame(rows, columns=cols).sort_values(
        "Current L2 Tickets Created By Champ", ascending=False
    ).reset_index(drop=True) if rows else pd.DataFrame(columns=cols)


def excel_bytes(raw, l2, tat_map):
    # Excel report functions use prepared/internal columns such as _Group,
    # _ClosedLike and _CurrentAgeHours. Prepare the full dump first.
    full = prepare_data(raw, tat_map)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    from openpyxl.utils import get_column_letter

    m=metrics(l2)
    buf=io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        overall_total = unique_ticket_count(full)
        overall_closed = unique_ticket_count(full[full["_ClosedLike"]])
        overall_open = unique_ticket_count(full[~full["_ClosedLike"]])

        overall_summary = pd.DataFrame([
            ["Overall Tickets Raised", overall_total],
            ["Overall Closed / Resolved", overall_closed],
            ["Overall Open / Pending", overall_open],
            ["Current L2 Tickets", unique_ticket_count(full[full["_IsL2"]])],
            ["Current L1 → L2 Handoff Proxy", l1_l2_proxy_count(full)],
        ], columns=["Metric", "Count"])
        overall_summary.to_excel(
            writer,
            index=False,
            sheet_name="Overall Summary"
        )

        summary=pd.DataFrame([
            ["Tickets Raised",m["raised"]],
            ["Closed / Resolved",m["closed"]],
            ["Open / Pending",m["open"]],
            ["Closed Within TAT",m["closed_within_tat"]],
            ["Closed Beyond TAT",m["closed_beyond_tat"]],
            ["TAT Eligible Closed / Resolved",m["tat_closed_eligible"]],
            ["TAT Compliance %",round(m["tat_compliance_pct"],1)],
            ["Open / Pending Beyond TAT",m["open_beyond_tat"]],
            ["TAT Mapped Tickets",m["tat_mapped"]],
            ["TAT Unmapped Tickets",m["tat_unmapped"]],
            ["Open / Pending >90 days",m["open_gt90"]],
            ["Average Resolution",format_hms(m["avg"])],
            ["Max Resolution",format_hms(m["max_resolution"])],
            ["90% Percentile Resolution",format_hms(m["p90"])],
            ["95% Percentile Resolution",format_hms(m["p95"])],
            ["99% Percentile Resolution",format_hms(m["p99"])],
            ["Max Current Open Age",format_hms(m["max_open_age"])],
        ],columns=["Metric","Value"])
        summary.to_excel(writer,index=False,sheet_name="Summary")

        cd_out=l2[[
            "Ticket ID","Subject","Status","Priority","Type","Agent","Group","CreatedBy",
            "Created time","Closed time","Last update time",
            "Resolution time (in hrs)","Resolution status",
            "Reason for pendency","Category","Sub-Category",
            "_CurrentAgeHours","_AgingBucket","_TATDays","_TATStatus"
        ]].copy()
        cd_out.rename(columns={
            "_CurrentAgeHours":"Current Age Hours",
            "_AgingBucket":"Current Aging Bucket",
            "_TATDays":"TAT Days",
            "_TATStatus":"TAT Status"
        },inplace=True)
        cd_out.to_excel(writer,index=False,sheet_name="L2 Raw + Analysis")

        aging_table(l2).to_excel(writer,index=False,sheet_name="Ticket Aging")
        open_reason_table(l2).to_excel(writer,index=False,sheet_name="Open Reasons")
        issue_breakup_table(l2, "Category").to_excel(writer,index=False,sheet_name="Category Breakup")
        issue_breakup_table(l2, "Sub-Category").to_excel(writer,index=False,sheet_name="Subcategory Breakup")
        subcategory_tat_performance_table(l2).to_excel(writer,index=False,sheet_name="Subcategory TAT Performance")
        category_subcategory_table(l2).to_excel(writer,index=False,sheet_name="Category + Subcategory")
        other_group_pending_table(full).to_excel(writer,index=False,sheet_name="Current Group Pending")
        other_group_open_breakup_table(full).to_excel(writer,index=False,sheet_name="Current Group Open Breakup")
        other_group_open_agent_breakup_table(full).to_excel(writer,index=False,sheet_name="Current Group Open Agent Breakup")
        other_group_open_category_breakup_table(full).to_excel(writer,index=False,sheet_name="Current Group Open Category")
        agent_summary_table(l2).to_excel(writer,index=False,sheet_name="Agent Summary")
        agent_ticket_detail_table(l2).to_excel(writer,index=False,sheet_name="Agent Ticket Detail")
        l1_agent_table(full).to_excel(writer,index=False,sheet_name="CD L1 Agent Summary")
        l1_ticket_detail_table(full).to_excel(writer,index=False,sheet_name="CD L1 Fresh Unassigned")
        fcr_l1_table(full).to_excel(writer,index=False,sheet_name="CD L1 FCR Proxy")
        l1_created_to_l2_table(full).to_excel(writer,index=False,sheet_name="L1 to L2 Handoff Proxy")
        other_group_agent_pending_table(full).to_excel(
            writer,index=False,sheet_name="Current Group Agent Pending"
        )
        other_group_ticket_detail_table(full).to_excel(
            writer,index=False,sheet_name="Current Group Ticket Detail"
        )
        group_aging_table(full).to_excel(writer,index=False,sheet_name="All Groups Aging")
        resolution_by_group(full).to_excel(writer,index=False,sheet_name="All Groups Resolution")

        # Monthly view for L2
        monthly=[]
        for key,g in l2.groupby("_MonthSort",sort=True):
            mm=metrics(g)
            monthly.append({
                "Month":g["_Month"].iloc[0],
                "Tickets Raised":mm["raised"],
                "Closed/Resolved":mm["closed"],
                "Open/Pending":mm["open"],
                "Closed Within TAT":mm["closed_within_tat"],
                "Closed Beyond TAT":mm["closed_beyond_tat"],
                "TAT Compliance %":round(mm["tat_compliance_pct"],1),
                "Avg Resolution":format_hms(mm["avg"]),
                "90%":format_hms(mm["p90"]),
                "95%":format_hms(mm["p95"]),
                "99%":format_hms(mm["p99"]),
            })
        pd.DataFrame(monthly).to_excel(writer,index=False,sheet_name="Monthly Trend")
        week_breakup_table(l2).to_excel(writer,index=False,sheet_name="Weekly Trend")

    wb=load_workbook(buf)
    header_fill=PatternFill("solid",fgColor="173A5E")
    white_font=Font(color="FFFFFF",bold=True)
    thin=Side(style="thin",color="D9E2EC")

    for ws in wb.worksheets:
        ws.freeze_panes="A2"
        ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]:
            cell.fill=header_fill
            cell.font=white_font
            cell.alignment=Alignment(horizontal="center",vertical="center")
        for row in ws.iter_rows():
            for c in row:
                c.border=Border(bottom=thin)
                c.alignment=Alignment(vertical="top")
        for col in range(1,ws.max_column+1):
            letter=get_column_letter(col)
            maxlen=0
            for cell in ws[letter][:200]:
                if cell.value is not None:
                    maxlen=max(maxlen,len(str(cell.value)))
            ws.column_dimensions[letter].width=min(max(maxlen+2,12),42)
        # Excel AutoFilter keeps the downloaded tables filterable without
        # requiring openpyxl TableStyleInfo compatibility.

    # Conditional formatting for useful SLA/aging columns.
    from openpyxl.formatting.rule import CellIsRule
    red=PatternFill("solid",fgColor="F4CCCC")
    green=PatternFill("solid",fgColor="D9EAD3")
    orange=PatternFill("solid",fgColor="FCE5CD")
    for wsname in ["L2 Raw + Analysis","All Groups Aging","Monthly Trend"]:
        ws=wb[wsname]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str):
                    if "Beyond TAT" in cell.value or ">90" in cell.value:
                        cell.fill=red
                    if "Within TAT" in cell.value:
                        cell.fill=green
    out=io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ============================================================
# HEADER
# ============================================================
st.markdown("""
<div class="hero">
  <h1>SOLV — L2 TICKET PERFORMANCE & TAT DASHBOARD</h1>
  <p>L2 bucket health, sub-category TAT, ticket aging, resolution performance and workload drivers.</p>
</div>
""", unsafe_allow_html=True)

uploaded=st.file_uploader(
    "Upload the SOLV ticket dump (CSV / Excel)",
    type=["csv","xlsx","xls"],
    help="The dashboard analyses the raw Group field and focuses the main KPIs on the configured L2 groups."
)

if uploaded is None:
    st.info("Upload the SOLV ticket dump to start the L2 analysis.")
    st.stop()

try:
    if uploaded.name.lower().endswith(".csv"):
        raw=pd.read_csv(uploaded,low_memory=False)
    else:
        raw=pd.read_excel(uploaded, engine="openpyxl" if uploaded.name.lower().endswith(".xlsx") else None)
except Exception as e:
    st.error(f"Could not read the file: {e}")
    st.stop()

missing=[c for c in REQUIRED if c not in raw.columns]
if missing:
    st.error("The file is missing required columns: " + ", ".join(missing))
    st.stop()

tat_info, tat_error = get_tat_data()

if tat_info is None:
    st.error("### Live TAT sheet could not be loaded")
    st.error(tat_error)
    st.info(
        "The dashboard is intentionally stopped here so it never falls back "
        "to an old/hard-coded TAT list."
    )
    st.stop()

tat_map = tat_info["map"]
df=prepare_data(raw, tat_map)

# ============================================================
# FILTERS
# ============================================================
st.markdown('<div class="section">FILTERS</div>',unsafe_allow_html=True)

# Month / Week / Day are all based on Created time.
f1,f2,f3,f4,f5=st.columns(5)

months=sorted(df["_MonthSort"].dropna().unique().tolist())
month_labels={"All":"All"}
for x in months:
    month_labels[x]=pd.Period(x).strftime("%b %Y")

with f1:
    selected_month=st.selectbox(
        "Month",
        ["All"]+[month_labels[x] for x in months]
    )

# Week options from the complete raw dump.
week_options=sorted(
    df.loc[df["_WeekStart"].notna(), "_WeekStart"]
    .drop_duplicates()
    .tolist()
)
week_labels=["All"]
week_map={}
for wk in week_options:
    label=(
        f"{wk.strftime('%d %b %Y')} - "
        f"{(wk + pd.Timedelta(days=6)).strftime('%d %b %Y')}"
    )
    week_labels.append(label)
    week_map[label]=wk

with f2:
    selected_week=st.selectbox("Week",week_labels)

# Day options from Created time.
day_values=sorted(
    df.loc[df["_CreatedDate"].notna(), "_CreatedDate"]
    .drop_duplicates()
    .tolist()
)
day_labels=["All"]
day_map={}
for d in day_values:
    label=pd.Timestamp(d).strftime("%d %b %Y")
    day_labels.append(label)
    day_map[label]=d

with f3:
    selected_day=st.selectbox("Day",day_labels)

with f4:
    status_options=["All"]+sorted(
        df["_Status"].dropna().unique().tolist()
    )
    selected_status=st.selectbox("Status",status_options)

with f5:
    category_options=["All"]+sorted(
        clean_text(df["Category"])
        .replace("","(blank)")
        .unique()
        .tolist()
    )
    selected_category=st.selectbox("Category",category_options)

# Apply the date filters to the complete population first.
filtered_all=df.copy()

if selected_month!="All":
    target=[x for x,label in month_labels.items() if label==selected_month]
    if target:
        filtered_all=filtered_all[
            filtered_all["_MonthSort"].eq(target[0])
        ]

if selected_week!="All":
    filtered_all=filtered_all[
        filtered_all["_WeekStart"].eq(week_map[selected_week])
    ]

if selected_day!="All":
    filtered_all=filtered_all[
        filtered_all["_CreatedDate"].eq(day_map[selected_day])
    ]

if selected_status!="All":
    filtered_all=filtered_all[
        filtered_all["_Status"].eq(selected_status)
    ]

if selected_category!="All":
    filtered_all=filtered_all[
        clean_text(filtered_all["Category"])
        .replace("","(blank)")
        .eq(selected_category)
    ]

# Main SOLV dashboard population = all configured L2 groups:
# Ops - L2, CD - Seller Support, Tech Support, Logistics, CD L2.
l2=filtered_all[
    filtered_all["_IsL2"]
].copy()

# L1 visibility is separate: CD L1 Team + No Group.
l1=filtered_all[
    filtered_all["_IsL1"]
].copy()

st.caption(
    f"Showing {unique_ticket_count(l2):,} current L2 tickets"
    + (f" | Month: {selected_month}" if selected_month!="All" else " | All available months")
    + (f" | Week: {selected_week}" if selected_week!="All" else "")
    + (f" | Day: {selected_day}" if selected_day!="All" else "")
    + (f" | Status: {selected_status}" if selected_status!="All" else "")
    + (f" | Category: {selected_category}" if selected_category!="All" else "")
)

if l2.empty:
    st.warning("No L2 tickets match the selected filters.")
    st.stop()


st.markdown('<div class="section">HOW TO READ THIS DASHBOARD</div>', unsafe_allow_html=True)
st.info(
    "1) Main L2 KPIs = current tickets in CD L2, Ops - L2, CD - Seller Support, Tech Support and Logistics. "
    "2) TAT is taken live from the Google Sheet by Sub-Category. "
    "3) Closed Within TAT / Closed Beyond TAT compare resolution time with that ticket's mapped TAT. "
    "4) Open Beyond TAT compares current age with that ticket's mapped TAT. "
    "5) Current Group tables always show the exact raw Group name. "
    "6) L1 → L2 is a CreatedBy proxy only; this dump does not contain historical transfer history. "
    "7) TAT Compliance % = Closed Within TAT ÷ (Closed Within TAT + Closed Beyond TAT)."
)

# ============================================================
# CD L1 — INTAKE / SAME-DAY / HANDOFF
# ============================================================
st.markdown(
    '<div class="section">CD L1 — INTAKE, SAME-DAY & L2 HANDOFF</div>',
    unsafe_allow_html=True
)
l1a = cd_l1_analysis(filtered_all)

q1,q2,q3,q4 = st.columns(4)
with q1:
    kpi("CD L1 TICKETS RAISED", f"{l1a['l1_raised']:,}", "blue",
        "Current Group = CD L1 Team / No Group")
with q2:
    kpi("L1 CLOSED SAME DAY", f"{l1a['l1_closed_same_day']:,}", "green",
        "Created date = Closed date")
with q3:
    kpi("CURRENTLY IN L2", f"{l1a['l1_current_l2']:,}", "orange",
        "Current L2 group; movement history is not available")
with q4:
    kpi("L1 → L2 HANDOFF PROXY", f"{l1a['l1_l2_handoff_proxy']:,}", "blue",
        "Current L2 + CreatedBy also seen in current L1; not historical movement")

q5,q6,q7,q8 = st.columns(4)
with q5:
    kpi("L1 OPEN / PENDING", f"{l1a['l1_open']:,}", "red")
with q6:
    kpi("L1 SAME-DAY PHONE PROXY", f"{l1a['same_day_phone_proxy']:,}", "green",
        "Phone + same-day close; not proven FCR")
with q7:
    kpi("CD L1 AGENT COUNT", f"{l1a['l1']['Agent'].nunique():,}", "blue")
with q8:
    kpi("L1 FRESH UNASSIGNED <48H", f"{l1a['l1_fresh_unassigned']:,}", "orange",
        "Open/pending + no agent + age <48h")

st.caption(
    "Important: the raw dump contains the current Group only. Therefore "
    "\"Currently in L2\" is a current-state count, not proof of historical "
    "movement from CD L1 to L2. A true FCR cannot be proven because this "
    "dump does not contain Call ID/contact history; the same-day Phone number "
    "is shown only as an operational proxy."
)

st.markdown(
    '<div class="section">CD L1 — AGENT-WISE WORKLOAD</div>',
    unsafe_allow_html=True
)
st.dataframe(
    l1_agent_table(filtered_all),
    use_container_width=True,
    hide_index=True
)

st.markdown(
    '<div class="section">CD L1 — FRESH UNASSIGNED TICKETS <48H</div>',
    unsafe_allow_html=True
)
st.dataframe(
    l1_ticket_detail_table(filtered_all),
    use_container_width=True,
    hide_index=True
)

# ============================================================
# OVERALL TICKET SUMMARY — FULL DUMP
# ============================================================
st.markdown(
    '<div class="section">OVERALL TICKET SUMMARY — FULL DUMP</div>',
    unsafe_allow_html=True
)

overall_total = unique_ticket_count(filtered_all)
overall_closed = unique_ticket_count(
    filtered_all[filtered_all["_ClosedLike"]]
)
overall_open = unique_ticket_count(
    filtered_all[~filtered_all["_ClosedLike"]]
)

overall_q1, overall_q2, overall_q3 = st.columns(3)

with overall_q1:
    kpi(
        "OVERALL TICKETS RAISED",
        f"{overall_total:,}",
        "blue",
        "All groups in the selected filter scope"
    )

with overall_q2:
    kpi(
        "OVERALL CLOSED / RESOLVED",
        f"{overall_closed:,}",
        "green",
        "All groups in the selected filter scope"
    )

with overall_q3:
    kpi(
        "OVERALL OPEN / PENDING",
        f"{overall_open:,}",
        "red",
        "All groups in the selected filter scope"
    )

overall_status = pd.DataFrame([
    {
        "Metric": "Overall Tickets Raised",
        "Count": overall_total,
        "% of Total": 100.0 if overall_total else 0.0,
    },
    {
        "Metric": "Overall Closed / Resolved",
        "Count": overall_closed,
        "% of Total": round(
            overall_closed / overall_total * 100, 1
        ) if overall_total else 0.0,
    },
    {
        "Metric": "Overall Open / Pending",
        "Count": overall_open,
        "% of Total": round(
            overall_open / overall_total * 100, 1
        ) if overall_total else 0.0,
    },
])

st.dataframe(
    overall_status,
    use_container_width=True,
    hide_index=True
)

# ============================================================
# L1 VISIBILITY
# ============================================================
st.markdown(
    '<div class="section">L1 VISIBILITY — CD L1 TEAM + NO GROUP</div>',
    unsafe_allow_html=True
)

l1_total = unique_ticket_count(l1)
l1_open = unique_ticket_count(l1[~l1["_ClosedLike"]])
l1_unassigned = l1[
    (~l1["_ClosedLike"]) &
    clean_text(l1["Agent"]).str.casefold().isin(["", "no agent", "not provided"])
]
l1_fresh_unassigned = l1_unassigned[
    l1_unassigned["_CurrentAgeHours"].notna() &
    (l1_unassigned["_CurrentAgeHours"] < 48)
]

l1q1,l1q2,l1q3,l1q4=st.columns(4)
with l1q1:
    kpi("L1 TICKETS", f"{l1_total:,}", "blue")
with l1q2:
    kpi("L1 OPEN / PENDING", f"{l1_open:,}", "red")
with l1q3:
    kpi("L1 FRESH & UNASSIGNED <48H", f"{unique_ticket_count(l1_fresh_unassigned):,}", "orange")
with l1q4:
    kpi("L1 → L2 HANDOFF PROXY", f"{l1_l2_proxy_count(filtered_all):,}", "green",
        "Current L2 + CreatedBy also seen in current L1")

st.markdown("**L1-created / current-L2 proxy by Created By**")
st.dataframe(
    l1_created_to_l2_table(filtered_all),
    use_container_width=True,
    hide_index=True
)

st.caption(
    "L1 = current Group 'CD L1 Team' or 'No Group'. "
    "L1 → L2 is shown as a current-state CreatedBy proxy because this raw dump "
    "does not contain Group movement history."
)

# ============================================================
# AGENT CROSS-GROUP DRILLDOWN
# ============================================================
st.markdown(
    '<div class="section">AGENT CROSS-GROUP WORKLOAD — ALL GROUPS</div>',
    unsafe_allow_html=True
)

all_agents = sorted(
    clean_text(filtered_all["Agent"])
    .replace("", "No Agent")
    .unique()
    .tolist()
)

if all_agents:
    selected_agent = st.selectbox(
        "Select agent / champ",
        all_agents
    )

    agent_all = group_set_for_agent(filtered_all, selected_agent)
    st.caption(
        f"{selected_agent} currently has "
        f"{unique_ticket_count(agent_all):,} tickets across "
        f"{agent_all['_Group'].nunique():,} group(s)."
    )

    agent_summary = agent_cross_group_summary(
        filtered_all,
        selected_agent
    )
    st.dataframe(
        agent_summary,
        use_container_width=True,
        hide_index=True
    )

    group_choices = ["All"] + sorted(agent_all["_Group"].unique().tolist())
    selected_agent_group = st.selectbox(
        "Expand / inspect group for selected agent",
        group_choices
    )

    agent_detail = agent_cross_group_tickets(
        filtered_all,
        selected_agent,
        selected_agent_group
    )

    with st.expander(
        f"View {selected_agent}'s ticket details"
    ):
        st.dataframe(
            agent_detail,
            use_container_width=True,
            hide_index=True
        )

# ============================================================
# SUMMARY
# ============================================================
m=metrics(l2)
st.markdown('<div class="section">SUMMARY — L2</div>',unsafe_allow_html=True)

r1=st.columns(4)
with r1[0]: kpi("TICKETS IN L2",f"{m['raised']:,}","blue","Total tickets raised in the configured L2 groups")
with r1[1]: kpi("CLOSED / RESOLVED",f"{m['closed']:,}","green","Closed or resolved tickets")
with r1[2]: kpi("OPEN / PENDING",f"{m['open']:,}","red","Current unresolved workload")
with r1[3]: kpi("TAT MAPPED / UNMAPPED",f"{m['tat_mapped']:,} / {m['tat_unmapped']:,}","dark","Tickets with / without a Sub-category TAT mapping")

r2=st.columns(4)
with r2[0]: kpi("CLOSED WITHIN TAT",f"{m['closed_within_tat']:,}","green","Closed within the mapped Sub-category TAT")
with r2[1]: kpi("CLOSED BEYOND TAT",f"{m['closed_beyond_tat']:,}","orange","Closed after the mapped Sub-category TAT")
with r2[2]: kpi("TAT COMPLIANCE %",f"{m['tat_compliance_pct']:.1f}%","blue","Closed within TAT ÷ TAT-eligible closed/resolved tickets")
with r2[3]: kpi("OPEN / PENDING BEYOND TAT",f"{m['open_beyond_tat']:,}","red","Current age is beyond the mapped Sub-category TAT")

with st.popover("🔎 VIEW CURRENT NON-L2 GROUP OPEN / PENDING BREAKUP"):
    st.markdown("### Current Non-L2 Group — Open / Pending Breakup")
    st.caption("This view shows the exact current Group names that are outside the five configured L2 groups (CD L2, Ops - L2, CD - Seller Support, Tech Support, Logistics), after the selected filters.")
    other_open_all = filtered_all[~filtered_all["_IsL2"] & ~filtered_all["_ClosedLike"]].copy()
    og_total = unique_ticket_count(other_open_all)
    og_open = unique_ticket_count(other_open_all[other_open_all["_Status"].eq("OPEN")])
    og_pending = unique_ticket_count(other_open_all[other_open_all["_Status"].eq("PENDING")])
    og_pending_brand = unique_ticket_count(other_open_all[other_open_all["_Status"].eq("PENDING FROM BRAND")])
    c1,c2,c3,c4 = st.columns(4)
    with c1: kpi("CURRENT NON-L2 UNRESOLVED", f"{og_total:,}", "red", "Open + Pending + other unresolved statuses")
    with c2: kpi("CURRENT NON-L2 OPEN", f"{og_open:,}", "orange")
    with c3: kpi("CURRENT NON-L2 PENDING", f"{og_pending:,}", "blue")
    with c4: kpi("PENDING FROM BRAND", f"{og_pending_brand:,}", "orange")
    st.markdown("**By Current Group**")
    st.dataframe(other_group_open_breakup_table(filtered_all), use_container_width=True, hide_index=True)
    st.markdown("**By Group + Agent**")
    st.dataframe(other_group_open_agent_breakup_table(filtered_all), use_container_width=True, hide_index=True)
    st.markdown("**By Group + Category + Sub-Category**")
    st.dataframe(other_group_open_category_breakup_table(filtered_all), use_container_width=True, hide_index=True)
    st.markdown("**Ticket-level detail**")
    st.dataframe(other_group_ticket_detail_table(filtered_all), use_container_width=True, hide_index=True)

r3=st.columns(4)
with r3[0]: kpi("AVG RESOLUTION",format_hms(m["avg"]),"blue","Valid closed/resolved tickets")
with r3[1]: kpi("MAX RESOLUTION",format_hms(m["max_resolution"]),"orange","Highest valid resolution time")
with r3[2]: kpi("MAX CURRENT AGE",format_hms(m["max_open_age"]),"red","Oldest open/pending ticket")
with r3[3]: kpi("VALID RESOLUTION TICKETS",f"{m['valid_resolution']:,}","dark","Used for average and percentiles")

st.caption(
    "TAT compliance is calculated on closed/resolved tickets with a mapped Sub-category TAT. "
    "TAT is mapped from the Sub-Category SLA sheet; tickets without a mapping are shown separately. Resolution-time metrics use valid closed/resolved tickets and the raw 'Resolution time (in hrs)' field."
)

# ============================================================
# PERCENTILES
# ============================================================
st.markdown('<div class="section">RESOLUTION TIME PERCENTILES</div>',unsafe_allow_html=True)
p1,p2,p3=st.columns(3)
with p1: kpi("90% PERCENTILE RESOLUTION HRS",format_hms(m["p90"]),"blue","PERCENTILE.INC-style 90th percentile")
with p2: kpi("95% PERCENTILE RESOLUTION HRS",format_hms(m["p95"]),"orange","PERCENTILE.INC-style 95th percentile")
with p3: kpi("99% PERCENTILE RESOLUTION HRS",format_hms(m["p99"]),"red","PERCENTILE.INC-style 99th percentile")

st.caption(
    "Percentiles are calculated from valid closed/resolved resolution times. "
    "They indicate the resolution-time point below which 90%, 95% or 99% of the valid tickets fall."
)

# ============================================================
# TAT MAPPING
# ============================================================
st.markdown('<div class="section">SUB-CATEGORY TAT MAPPING</div>',unsafe_allow_html=True)
tat_map_df = tat_info["table"].copy().sort_values("Subcategory").reset_index(drop=True)
st.dataframe(tat_map_df, use_container_width=True, hide_index=True)
st.caption(
    f"Live source: Google Sheet | {tat_info['rows']:,} TAT rows | "
    f"{tat_info['mapped_rows']:,} sub-categories with numeric TAT | "
    f"{tat_info['rows'] - tat_info['mapped_rows']:,} rows with NA/non-numeric TAT."
)
if tat_info["duplicate_subcategories"]:
    st.warning(
        f"The live TAT sheet contains {tat_info['duplicate_subcategories']} "
        "duplicate sub-category key(s). The last populated TAT is used for ticket matching."
    )
st.caption(
    f"Ticket mapping in the selected L2 population: {m['tat_mapped']:,} mapped tickets and "
    f"{m['tat_unmapped']:,} unmapped tickets. No fixed 48-hour SLA is used."
)

st.markdown('<div class="section">SUB-CATEGORY-WISE TAT PERFORMANCE</div>',unsafe_allow_html=True)
st.dataframe(
    subcategory_tat_performance_table(l2),
    use_container_width=True,
    hide_index=True
)

# ============================================================
# OPEN REASONS
# ============================================================
st.markdown('<div class="section">OPEN / PENDING REASONS</div>',unsafe_allow_html=True)
orows=open_reason_table(l2)
st.dataframe(orows,use_container_width=True,hide_index=True)

# ============================================================
# CATEGORY / SUB-CATEGORY ANALYSIS
# ============================================================
st.markdown('<div class="section">CATEGORY-WISE BREAKUP</div>',unsafe_allow_html=True)
cat_tbl = issue_breakup_table(l2, "Category")
st.dataframe(cat_tbl, use_container_width=True, hide_index=True)

st.markdown('<div class="section">SUB-CATEGORY-WISE BREAKUP</div>',unsafe_allow_html=True)
sub_tbl = issue_breakup_table(l2, "Sub-Category")
st.dataframe(sub_tbl, use_container_width=True, hide_index=True)

st.markdown('<div class="section">CATEGORY + SUB-CATEGORY CONTRIBUTION</div>',unsafe_allow_html=True)
cs_tbl = category_subcategory_table(l2)
st.dataframe(cs_tbl, use_container_width=True, hide_index=True)

# ============================================================
# OTHER GROUP / PENDING VIEW
# ============================================================
st.markdown('<div class="section">CURRENT NON-L2 GROUP TICKETS</div>',unsafe_allow_html=True)
st.caption(
    "Based only on the raw Group column. L2 remains the main dashboard "
    "population. This separate view shows tickets currently sitting in any "
    "group other than L2, including open/pending aging."
)
other_tbl = other_group_pending_table(filtered_all)
st.dataframe(other_tbl, use_container_width=True, hide_index=True)

# ============================================================
# AGENT / TICKET OWNERSHIP VIEW
# ============================================================
st.markdown('<div class="section">AGENT-WISE TICKET OWNERSHIP & RESOLUTION</div>',unsafe_allow_html=True)
st.caption(
    "Shows which agent is handling the selected L2 tickets, how many they "
    "have, their open/closed workload, aging and resolution time."
)
agent_sum = agent_summary_table(l2)
st.dataframe(agent_sum, use_container_width=True, hide_index=True)

st.markdown('<div class="section">AGENT / TICKET-LEVEL DETAIL</div>',unsafe_allow_html=True)
st.caption(
    "Ticket-level view: Agent, Group, Ticket ID, Created Date, Last Updated Date, "
    "Category, Sub-Category, resolution/aging time and Sub-category TAT status."
)
agent_detail = agent_ticket_detail_table(l2)
st.dataframe(agent_detail, use_container_width=True, hide_index=True)

# ============================================================
# OTHER GROUP — AGENT LEVEL PENDING WORKLOAD
# ============================================================
st.markdown(
    '<div class="section">CURRENT NON-L2 GROUP — AGENT-WISE PENDING TICKETS</div>',
    unsafe_allow_html=True
)
st.caption(
    "Separate from the L2 KPIs. This view shows currently open/pending "
    "tickets whose raw Group is not L2, grouped by agent and destination group."
)
other_agent_tbl = other_group_agent_pending_table(filtered_all)
st.dataframe(other_agent_tbl, use_container_width=True, hide_index=True)

st.markdown(
    '<div class="section">CURRENT NON-L2 GROUP — PENDING TICKET DETAIL</div>',
    unsafe_allow_html=True
)
other_detail_tbl = other_group_ticket_detail_table(filtered_all)
st.dataframe(other_detail_tbl, use_container_width=True, hide_index=True)

# ============================================================
# TICKET AGING
# ============================================================
st.markdown('<div class="section">CURRENT TICKET AGING</div>',unsafe_allow_html=True)
st.caption("For open/pending tickets, aging is measured from Created time to the current dashboard refresh time.")

ag=aging_table(l2)
st.dataframe(ag,use_container_width=True,hide_index=True)

# ============================================================
# GROUP TICKETING AGING
# ============================================================
st.markdown('<div class="section">GROUP TICKETING AGING — FULL DUMP</div>',unsafe_allow_html=True)
st.caption("This view uses the raw Group field across the complete dump, so you can see where the overall workload sits. Main KPIs above remain L2 only.")
ga=group_aging_table(df)
st.dataframe(ga,use_container_width=True,hide_index=True)

# ============================================================
# RESOLUTION BY GROUP
# ============================================================
st.markdown('<div class="section">GROUP RESOLUTION PERFORMANCE — FULL DUMP</div>',unsafe_allow_html=True)
gr=resolution_by_group(df)
st.dataframe(gr,use_container_width=True,hide_index=True)

# ============================================================
# WEEKLY TREND
# ============================================================
st.markdown('<div class="section">WEEK-WISE L2 BREAKUP</div>',unsafe_allow_html=True)
weekly_df = week_breakup_table(l2)
st.dataframe(weekly_df, use_container_width=True, hide_index=True)

# ============================================================
# MONTHLY TREND
# ============================================================
st.markdown('<div class="section">MONTH-WISE L2 BREAKUP</div>',unsafe_allow_html=True)
monthly=[]
for key,g in l2.groupby("_MonthSort",sort=True):
    mm=metrics(g)
    monthly.append({
        "Month":g["_Month"].iloc[0],
        "Raised":mm["raised"],
        "Closed / Resolved":mm["closed"],
        "Open / Pending":mm["open"],
        "Closed Within TAT":mm["closed_within_tat"],
        "Closed Beyond TAT":mm["closed_beyond_tat"],
        "TAT Compliance %":round(mm["tat_compliance_pct"],1),
        "Avg Resolution":format_hms(mm["avg"]),
        "90%":format_hms(mm["p90"]),
        "95%":format_hms(mm["p95"]),
        "99%":format_hms(mm["p99"]),
    })
monthly_df=pd.DataFrame(monthly)
st.dataframe(monthly_df,use_container_width=True,hide_index=True)

# ============================================================
# DOWNLOAD
# ============================================================
st.markdown('<div class="section">DOWNLOAD REPORT</div>',unsafe_allow_html=True)
st.caption(
    "The Excel report contains L2 raw data with analysis columns, "
    "category/sub-category breakups, sub-category TAT mapping/performance, TAT status, open reasons, ticket aging, other-group "
    "pending tickets, other-group open breakup, group aging, group resolution, weekly and monthly trends. "
    "Every main table has Excel filters enabled."
)

try:
    xlsx=excel_bytes(raw,l2,tat_map)
    st.download_button(
        "⬇️ Download L2 Analysis Excel",
        data=xlsx,
        file_name="SOLV_CD_L2_Ticket_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
except Exception as e:
    st.error(f"Could not create the Excel report: {e}")
